import openpyxl

def get_maze_moves_from_excel(sheet):
    moves_matrix = []

    for row in sheet.iter_rows(values_only = True):
        row_moves = []

        for cell in row:
            row_moves.append(cell)

        moves_matrix.append(row_moves)

    return moves_matrix

workbook = openpyxl.load_workbook("MazeList.xlsx", data_only = True)
sheet_list = workbook.sheetnames

for sheet in sheet_list:
    # Start element
    print(f"// {sheet}")
    print("{")

    # Start indicator list
    print("\t{")

    # Empty indicator data
    print("\t\t{ 0, 0 }, { 0, 0 },")

    # End indicator list
    print("\t},")

    # Start sheet data
    print("\t{{")

    # Print sheet data
    maze_moves = get_maze_moves_from_excel(workbook[sheet])
    if maze_moves:
        for row in maze_moves:
            print("\t\t" + str(row).replace("[", "{ ").replace("]", " }").replace("'", "\""), end = ",\n")

    # End sheet data
    print("\t}},")

    # End element
    print("},")
